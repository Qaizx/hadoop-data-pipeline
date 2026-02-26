import os
from dotenv import load_dotenv
import pandas as pd
from openai import OpenAI
from openpyxl import load_workbook

load_dotenv()  # โหลดตัวแปรจาก .env

client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

def analyze_excel_structure(excel_path: str, sheet_name: str) -> str:
    """อ่านโครงสร้าง Excel ส่งให้ GPT วิเคราะห์"""
    
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]
    
    # อ่าน 10 rows แรก
    preview = []
    for row_idx, row in enumerate(ws.iter_rows(max_row=10), start=1):
        row_data = []
        for cell in row[:20]:
            row_data.append(f"{cell.coordinate}:{cell.value}")
        preview.append(f"Row {row_idx}: {row_data}")
    
    # รวม merged cells info
    merged_info = [str(m) for m in list(ws.merged_cells.ranges)[:30]]
    
    return f"""
Sheet name: {ws.title}
Total rows: {ws.max_row}
Total columns: {ws.max_column}

First 10 rows preview:
{chr(10).join(preview)}

Merged Cells (first 30):
{merged_info}
"""

def ask_gpt_for_conversion_code(excel_structure: str, sheet_name: str) -> str:
    """ให้ GPT เขียนโค้ดแปลง Excel เป็น CSV"""
    
    response = client.chat.completions.create(
        model="gpt-4o",
        messages=[
            {
                "role": "system",
                "content": f"""คุณเป็น Python expert ที่เชี่ยวชาญ openpyxl และ pandas

ให้วิเคราะห์โครงสร้าง Excel แล้วเขียน function ชื่อ convert_excel(excel_path) ที่:
1. อ่านเฉพาะ sheet "{sheet_name}"
2. Handle merged cells โดย fill ค่าลงทุก cell ที่ถูก merge
3. ถ้ามี multi-level header ให้รวมเป็น column name เดียว คั่นด้วย " > "
4. Return pandas DataFrame

ตอบเฉพาะ Python code เท่านั้น ไม่ต้องอธิบาย ไม่ต้องใส่ markdown"""
            },
            {
                "role": "user", 
                "content": f"วิเคราะห์และเขียนโค้ดแปลง Excel นี้:\n{excel_structure}"
            }
        ]
    )
    
    code = response.choices[0].message.content.strip()
    
    # ลบ markdown code block
    if code.startswith("```"):
        lines = code.split("\n")
        lines = [l for l in lines if not l.startswith("```")]
        code = "\n".join(lines).strip()
    
    return code

import pandas as pd
from openpyxl import load_workbook

def convert_excel(excel_path: str, sheet_name: str) -> pd.DataFrame:
    """แปลง Excel ที่มี merged cells เป็น DataFrame"""
    
    wb = load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]
    
    # เก็บค่าจาก merged cells
    merged_values = {}
    for merged_range in ws.merged_cells.ranges:
        top_left_value = ws.cell(merged_range.min_row, merged_range.min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_values[(row, col)] = top_left_value
    
    # อ่านข้อมูลทั้งหมด
    data = []
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        row_data = []
        for col_idx, cell in enumerate(row, start=1):
            if (row_idx, col_idx) in merged_values:
                row_data.append(merged_values[(row_idx, col_idx)])
            else:
                row_data.append(cell.value)
        data.append(row_data)
    
    # สร้าง header จาก 4 rows แรก - รวมทุก level
    header_rows = data[:4]
    
    combined_headers = []
    for col_idx in range(len(header_rows[0])):
        parts = []
        for row in header_rows:
            if col_idx < len(row) and row[col_idx] is not None:
                val = str(row[col_idx]).strip().replace('\n', ' ')
                if val and val not in parts:
                    parts.append(val)
        col_name = "_".join(parts) if parts else f"col_{col_idx}"
        combined_headers.append(col_name)
    
    # ทำให้ column names unique
    seen = {}
    unique_headers = []
    for h in combined_headers:
        if h in seen:
            seen[h] += 1
            unique_headers.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 0
            unique_headers.append(h)
    
    df = pd.DataFrame(data[4:], columns=unique_headers)
    
    # ===== Clean Data =====
    df = df.dropna(axis=1, how='all')
    df = df.loc[:, ~df.columns.str.startswith('col_')]
    
    # Rename columns เป็นภาษาอังกฤษ (ครบทุก column)
    rename_map = {
        'ว/ด/ป': 'month',
        'จำนวนเงินรวม': 'total_amount',
        'รายละเอียด': 'type',
        
        # กองทุนทั่วไป งานบริหารทั่วไป
        'กองทุนทั่วไป งานบริหารทั่วไป_เงินอุดหนุน_wifi': 'general_subsidy_wifi',
        'ค่าตอบแทน_งบประจำ': 'general_compensation',
        'ค่าใช้สอย_งบประจำ': 'general_service',
        'ค่าวัสดุ_งบประจำ': 'general_material',
        'ค่าสาธารณูปโภค': 'general_utility',
        'เงินอุดหนุน_(สวัสดิการและสุขภาพบุคลากร)': 'general_welfare',
        'เงินอุดหนุน_MS 365': 'general_ms365',
        
        # กองทุนเพื่อการศึกษา
        'กองทุนเพื่อการศึกษา_งานบริการคอมพิวเตอร์ทางวิชาการ_เงินเดือน_พนง.เงินรายได้(S)': 'education_salary_income',
        'พนง.เงินแผ่นดิน': 'education_salary_gov',
        
        # กองทุนสินทรัพย์ถาวร
        'กองทุนสินทรัพย์ถาวร_งานบริการคอมพิวเตอร์ทางวิชาการ_ค่าครุภัณฑ์_(วงเงินไม่เกิน 1ล้าน)  งบประจำ': 'asset_equipment_under1m',
        'ค่าครุภัณฑ์_(วงเงินเกิน 1ล้าน)  งบประจำ': 'asset_equipment_over1m',
        'กองทุนสินทรัพย์ถาวร-ที่ดินและสิ่งก่อสร้าง': 'asset_land_building',
        'ครุภัณฑ์_Firewall': 'asset_firewall',
        'เงินอุดหนุน_SIEM': 'asset_siem',
        'เงินอุดหนุน_data center': 'asset_datacenter',
        'เงินอุดหนุน_wifi satit': 'asset_wifi_satit',
        
        # กองทุนเพื่อการวิจัย
        'กองทุนเพื่อการวิจัย_งานบริหารการวิจัย_วิจัยบุคลากร_เงินอุดหนุนทั่วไป': 'research_subsidy',
        
        # กองทุนสำรอง
        'กองทุนสำรอง_งานบริหารทั่วไป-รายจ่ายอื่น_สำรองจ่าย': 'reserve_expense',
        'สบทบกองทุนพัฒนา': 'reserve_dev_fund',
        'สมทบกองทุนพัฒนาบุคลากร มช': 'reserve_staff_fund',
        
        # กองทุนพัฒนาบุคลากร
        'กองทุนพัฒนาบุคลากร_งานสนับสนุนการจัดการศึกษา_อุดหนุนเฉพาะกิจ': 'development_specific',
        
        # กองทุนทำนุบำรุงศิลปะ
        'กองทุนทำนุบำรุงศิลปะ_งานทำนุฯ_อุดหนุนทั่วไป': 'culture_subsidy',
        
        # IT projects
        'Wifi@Jumboplus': 'it_wifi_jumbo',
        'Firewall': 'it_firewall',
        'CMU Cloud': 'it_cmu_cloud',
        'SiEM': 'it_siem',
        'Digital Health': 'it_digital_health',
        'ระบบการขอเข้าทำประโยชน์': 'it_access_system',
        'UPS': 'it_ups',
        'เช่าUPS+ดูแลwifi': 'it_ups_rental',
        'Uplift': 'it_uplift',
        'Open data': 'it_open_data',
    }
    
    df = df.rename(columns=rename_map)
    
    # แปลงชื่อเดือน
    month_map = {
        'ตค66': '2023-10', 'พย66': '2023-11', 'ธค66': '2023-12',
        'มค67': '2024-01', 'กพ67': '2024-02', 'มีค67': '2024-03',
        'เมย67': '2024-04', 'พค67': '2024-05', 'มิย67': '2024-06',
        'กค67': '2024-07', 'สค67': '2024-08', 'กย67': '2024-09',
    }
    df['month'] = df['month'].map(month_map).fillna(df['month'])
    
    # แปลง type
    type_map = {
        'ยอดงบประมาณ': 'budget',
        'ยอดใช้ไป': 'spent',
        'ยอดคงเหลือ': 'remaining',
    }
    df['type'] = df['type'].map(type_map).fillna(df['type'])
    
    # เอาเฉพาะ row ที่มี type ที่ต้องการ
    valid_types = ['budget', 'spent', 'remaining']
    df = df[df['type'].isin(valid_types)].copy()
    
    # Fill month ลงมา
    df['month'] = df['month'].ffill()
    
    df = df.reset_index(drop=True)
    
    # แสดง columns ที่ยังไม่ได้ rename
    thai_cols = [c for c in df.columns if any('\u0e00' <= ch <= '\u0e7f' for ch in c)]
    if thai_cols:
        print(f"⚠️ Columns ยังเป็นภาษาไทย: {thai_cols}")
    
    return df


# ============ Main ============
if __name__ == "__main__":
    excel_path = "data/คุมฎีกาปี2567-รายได้.xlsx"
    sheet_name = "สรุปรายเดือน 67"
    csv_path = "data/finance_2567_wide_en.csv"
    
    print(f"📊 Converting: {excel_path}")
    
    df = convert_excel(excel_path, sheet_name)
    
    df.to_csv(csv_path, index=False, encoding='utf-8-sig')
    
    print(f"✅ Saved to {csv_path}")
    print(f"   Rows: {len(df)}, Columns: {len(df.columns)}")
    
    print("\n📋 Columns:")
    for i, col in enumerate(df.columns, 1):
        print(f"  {i}. {col}")