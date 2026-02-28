import pandas as pd
from openpyxl import load_workbook
from openai import OpenAI
import json
import os
from dotenv import load_dotenv

load_dotenv()  # ‡πÇ‡∏´‡∏•‡∏î‡∏ï‡∏±‡∏ß‡πÅ‡∏õ‡∏£‡∏à‡∏≤‡∏Å .env

client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))


def get_column_mapping_from_gpt(columns: list) -> dict:
    """‡πÉ‡∏´‡πâ GPT ‡∏™‡∏£‡πâ‡∏≤‡∏á mapping ‡∏à‡∏≤‡∏Å‡∏ä‡∏∑‡πà‡∏≠ column ‡πÑ‡∏ó‡∏¢‡πÄ‡∏õ‡πá‡∏ô‡∏≠‡∏±‡∏á‡∏Å‡∏§‡∏©"""
    
    response = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {
                "role": "system",
                "content": """‡∏Ñ‡∏∏‡∏ì‡πÄ‡∏õ‡πá‡∏ô expert ‡πÉ‡∏ô‡∏Å‡∏≤‡∏£‡πÅ‡∏õ‡∏•‡∏á‡∏ä‡∏∑‡πà‡∏≠ column ‡∏à‡∏≤‡∏Å‡∏†‡∏≤‡∏©‡∏≤‡πÑ‡∏ó‡∏¢‡πÄ‡∏õ‡πá‡∏ô‡∏†‡∏≤‡∏©‡∏≤‡∏≠‡∏±‡∏á‡∏Å‡∏§‡∏©

‡∏Å‡∏é:
1. ‡πÉ‡∏ä‡πâ snake_case ‡πÄ‡∏ó‡πà‡∏≤‡∏ô‡∏±‡πâ‡∏ô (‡πÄ‡∏ä‡πà‡∏ô general_expense, asset_firewall)
2. ‡∏ä‡∏∑‡πà‡∏≠‡∏ï‡πâ‡∏≠‡∏á‡∏™‡∏±‡πâ‡∏ô ‡∏Å‡∏£‡∏∞‡∏ä‡∏±‡∏ö ‡πÅ‡∏ï‡πà‡πÄ‡∏Ç‡πâ‡∏≤‡πÉ‡∏à‡∏á‡πà‡∏≤‡∏¢
3. ‡∏ñ‡πâ‡∏≤‡πÄ‡∏õ‡πá‡∏ô‡∏†‡∏≤‡∏©‡∏≤‡∏≠‡∏±‡∏á‡∏Å‡∏§‡∏©‡∏≠‡∏¢‡∏π‡πà‡πÅ‡∏•‡πâ‡∏ß ‡πÉ‡∏´‡πâ‡πÅ‡∏õ‡∏•‡∏á‡πÄ‡∏õ‡πá‡∏ô snake_case
4. ‡∏ï‡∏≠‡∏ö‡πÄ‡∏õ‡πá‡∏ô JSON object ‡πÄ‡∏ó‡πà‡∏≤‡∏ô‡∏±‡πâ‡∏ô ‡πÑ‡∏°‡πà‡∏ï‡πâ‡∏≠‡∏á‡∏≠‡∏ò‡∏¥‡∏ö‡∏≤‡∏¢
5. format: {"‡∏ä‡∏∑‡πà‡∏≠‡πÄ‡∏î‡∏¥‡∏°": "‡∏ä‡∏∑‡πà‡∏≠‡πÉ‡∏´‡∏°‡πà", ...}"""
            },
            {
                "role": "user",
                "content": f"‡πÅ‡∏õ‡∏•‡∏á‡∏ä‡∏∑‡πà‡∏≠ columns ‡πÄ‡∏´‡∏•‡πà‡∏≤‡∏ô‡∏µ‡πâ‡πÄ‡∏õ‡πá‡∏ô‡∏†‡∏≤‡∏©‡∏≤‡∏≠‡∏±‡∏á‡∏Å‡∏§‡∏©:\n{json.dumps(columns, ensure_ascii=False)}"
            }
        ]
    )
    
    result = response.choices[0].message.content.strip()
    if result.startswith("```"):
        lines = result.split("\n")
        lines = [l for l in lines if not l.startswith("```")]
        result = "\n".join(lines).strip()
    
    return json.loads(result)


def get_data_mapping_from_gpt(values: list) -> dict:
    """‡πÉ‡∏´‡πâ GPT ‡∏™‡∏£‡πâ‡∏≤‡∏á mapping ‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡πÉ‡∏ô‡πÅ‡∏ï‡πà‡∏•‡∏∞ cell"""
    
    response = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {
                "role": "system",
                "content": """‡πÅ‡∏õ‡∏•‡∏á‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏†‡∏≤‡∏©‡∏≤‡πÑ‡∏ó‡∏¢‡πÄ‡∏õ‡πá‡∏ô‡∏†‡∏≤‡∏©‡∏≤‡∏≠‡∏±‡∏á‡∏Å‡∏§‡∏©

‡∏Å‡∏é:
1. ‡πÄ‡∏î‡∏∑‡∏≠‡∏ô‡πÑ‡∏ó‡∏¢ -> YYYY-MM format (‡∏õ‡∏µ 66=2023, 67=2024, 68=2025)
   - ‡∏ï‡∏Ñ66 -> 2023-10, ‡∏û‡∏¢66 -> 2023-11, ‡∏ò‡∏Ñ66 -> 2023-12
   - ‡∏°‡∏Ñ67 -> 2024-01, ‡∏Å‡∏û67 -> 2024-02, ‡∏°‡∏µ‡∏Ñ67 -> 2024-03
   - ‡πÄ‡∏°‡∏¢67 -> 2024-04, ‡∏û‡∏Ñ67 -> 2024-05, ‡∏°‡∏¥‡∏¢67 -> 2024-06
   - ‡∏Å‡∏Ñ67 -> 2024-07, ‡∏™‡∏Ñ67 -> 2024-08, ‡∏Å‡∏¢67 -> 2024-09

2. ‡∏õ‡∏£‡∏∞‡πÄ‡∏†‡∏ó‡∏£‡∏≤‡∏¢‡∏Å‡∏≤‡∏£ -> English
   - ‡∏¢‡∏≠‡∏î‡∏á‡∏ö‡∏õ‡∏£‡∏∞‡∏°‡∏≤‡∏ì -> budget
   - ‡∏¢‡∏≠‡∏î‡πÉ‡∏ä‡πâ‡πÑ‡∏õ -> spent
   - ‡∏¢‡∏≠‡∏î‡∏Ñ‡∏á‡πÄ‡∏´‡∏•‡∏∑‡∏≠ -> remaining

3. ‡∏Ç‡πâ‡∏≠‡∏Ñ‡∏ß‡∏≤‡∏°‡∏≠‡∏∑‡πà‡∏ô‡πÜ -> ‡πÅ‡∏õ‡∏•‡πÄ‡∏õ‡πá‡∏ô‡∏†‡∏≤‡∏©‡∏≤‡∏≠‡∏±‡∏á‡∏Å‡∏§‡∏©‡πÅ‡∏ö‡∏ö‡∏™‡∏±‡πâ‡∏ô‡∏Å‡∏£‡∏∞‡∏ä‡∏±‡∏ö

4. ‡∏ï‡∏≠‡∏ö‡πÄ‡∏õ‡πá‡∏ô JSON object ‡πÄ‡∏ó‡πà‡∏≤‡∏ô‡∏±‡πâ‡∏ô: {"‡∏Ñ‡πà‡∏≤‡πÄ‡∏î‡∏¥‡∏°": "‡∏Ñ‡πà‡∏≤‡πÉ‡∏´‡∏°‡πà", ...}
5. ‡∏ñ‡πâ‡∏≤‡πÄ‡∏õ‡πá‡∏ô‡∏ï‡∏±‡∏ß‡πÄ‡∏•‡∏Ç‡∏´‡∏£‡∏∑‡∏≠‡∏†‡∏≤‡∏©‡∏≤‡∏≠‡∏±‡∏á‡∏Å‡∏§‡∏©‡∏≠‡∏¢‡∏π‡πà‡πÅ‡∏•‡πâ‡∏ß ‡πÑ‡∏°‡πà‡∏ï‡πâ‡∏≠‡∏á‡πÉ‡∏™‡πà‡πÉ‡∏ô mapping"""
            },
            {
                "role": "user",
                "content": f"‡πÅ‡∏õ‡∏•‡∏á‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡πÄ‡∏´‡∏•‡πà‡∏≤‡∏ô‡∏µ‡πâ:\n{json.dumps(values, ensure_ascii=False)}"
            }
        ]
    )
    
    result = response.choices[0].message.content.strip()
    if result.startswith("```"):
        lines = result.split("\n")
        lines = [l for l in lines if not l.startswith("```")]
        result = "\n".join(lines).strip()
    
    return json.loads(result)


def analyze_dataframe_structure(df: pd.DataFrame) -> dict:
    """‡πÉ‡∏´‡πâ GPT ‡∏ß‡∏¥‡πÄ‡∏Ñ‡∏£‡∏≤‡∏∞‡∏´‡πå‡πÇ‡∏Ñ‡∏£‡∏á‡∏™‡∏£‡πâ‡∏≤‡∏á DataFrame ‡πÅ‡∏•‡∏∞‡∏ö‡∏≠‡∏Å‡∏ß‡πà‡∏≤ column ‡πÑ‡∏´‡∏ô‡∏Ñ‡∏∑‡∏≠‡∏≠‡∏∞‡πÑ‡∏£"""
    
    # ‡∏™‡∏£‡πâ‡∏≤‡∏á sample data ‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö‡πÉ‡∏´‡πâ GPT ‡∏ß‡∏¥‡πÄ‡∏Ñ‡∏£‡∏≤‡∏∞‡∏´‡πå
    sample_data = {}
    for col in df.columns[:10]:  # ‡πÄ‡∏≠‡∏≤‡πÅ‡∏Ñ‡πà 10 columns ‡πÅ‡∏£‡∏Å
        unique_vals = df[col].dropna().unique()[:5].tolist()  # ‡πÄ‡∏≠‡∏≤‡πÅ‡∏Ñ‡πà 5 values
        sample_data[col] = unique_vals
    
    response = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {
                "role": "system",
                "content": """‡∏ß‡∏¥‡πÄ‡∏Ñ‡∏£‡∏≤‡∏∞‡∏´‡πå‡πÇ‡∏Ñ‡∏£‡∏á‡∏™‡∏£‡πâ‡∏≤‡∏á DataFrame ‡πÅ‡∏•‡∏∞‡∏£‡∏∞‡∏ö‡∏∏:
1. month_column: ‡∏ä‡∏∑‡πà‡∏≠ column ‡∏ó‡∏µ‡πà‡πÄ‡∏Å‡πá‡∏ö‡πÄ‡∏î‡∏∑‡∏≠‡∏ô (‡∏ñ‡πâ‡∏≤‡∏°‡∏µ)
2. type_column: ‡∏ä‡∏∑‡πà‡∏≠ column ‡∏ó‡∏µ‡πà‡πÄ‡∏Å‡πá‡∏ö‡∏õ‡∏£‡∏∞‡πÄ‡∏†‡∏ó (budget/spent/remaining) (‡∏ñ‡πâ‡∏≤‡∏°‡∏µ)
3. type_values: ‡∏Ñ‡πà‡∏≤‡∏ó‡∏µ‡πà‡∏´‡∏°‡∏≤‡∏¢‡∏ñ‡∏∂‡∏á budget, spent, remaining

‡∏ï‡∏≠‡∏ö‡πÄ‡∏õ‡πá‡∏ô JSON:
{
    "month_column": "‡∏ä‡∏∑‡πà‡∏≠ column ‡∏´‡∏£‡∏∑‡∏≠ null",
    "type_column": "‡∏ä‡∏∑‡πà‡∏≠ column ‡∏´‡∏£‡∏∑‡∏≠ null",
    "type_values": {
        "budget": "‡∏Ñ‡πà‡∏≤‡∏ó‡∏µ‡πà‡∏´‡∏°‡∏≤‡∏¢‡∏ñ‡∏∂‡∏á budget ‡∏´‡∏£‡∏∑‡∏≠ null",
        "spent": "‡∏Ñ‡πà‡∏≤‡∏ó‡∏µ‡πà‡∏´‡∏°‡∏≤‡∏¢‡∏ñ‡∏∂‡∏á spent ‡∏´‡∏£‡∏∑‡∏≠ null",
        "remaining": "‡∏Ñ‡πà‡∏≤‡∏ó‡∏µ‡πà‡∏´‡∏°‡∏≤‡∏¢‡∏ñ‡∏∂‡∏á remaining ‡∏´‡∏£‡∏∑‡∏≠ null"
    }
}"""
            },
            {
                "role": "user",
                "content": f"‡∏ß‡∏¥‡πÄ‡∏Ñ‡∏£‡∏≤‡∏∞‡∏´‡πå DataFrame ‡∏ô‡∏µ‡πâ:\nColumns: {df.columns.tolist()}\nSample data: {json.dumps(sample_data, ensure_ascii=False, default=str)}"
            }
        ]
    )
    
    result = response.choices[0].message.content.strip()
    if result.startswith("```"):
        lines = result.split("\n")
        lines = [l for l in lines if not l.startswith("```")]
        result = "\n".join(lines).strip()
    
    return json.loads(result)


def convert_excel(excel_path: str, sheet_name: str) -> pd.DataFrame:
    """‡πÅ‡∏õ‡∏•‡∏á Excel ‡∏ó‡∏µ‡πà‡∏°‡∏µ merged cells ‡πÄ‡∏õ‡πá‡∏ô DataFrame"""
    
    wb = load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]
    
    # ‡πÄ‡∏Å‡πá‡∏ö‡∏Ñ‡πà‡∏≤‡∏à‡∏≤‡∏Å merged cells
    merged_values = {}
    for merged_range in ws.merged_cells.ranges:
        top_left_value = ws.cell(merged_range.min_row, merged_range.min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_values[(row, col)] = top_left_value
    
    # ‡∏≠‡πà‡∏≤‡∏ô‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡∏ó‡∏±‡πâ‡∏á‡∏´‡∏°‡∏î
    data = []
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        row_data = []
        for col_idx, cell in enumerate(row, start=1):
            if (row_idx, col_idx) in merged_values:
                row_data.append(merged_values[(row_idx, col_idx)])
            else:
                row_data.append(cell.value)
        data.append(row_data)
    
    # ‡∏™‡∏£‡πâ‡∏≤‡∏á header ‡∏à‡∏≤‡∏Å 4 rows ‡πÅ‡∏£‡∏Å
    header_rows = data[:4]
    
    combined_headers = []
    for col_idx in range(len(header_rows[0])):
        parts = []
        for row in header_rows:
            if col_idx < len(row) and row[col_idx] is not None:
                val = str(row[col_idx]).strip().replace('\n', ' ')
                if val and val not in parts:
                    parts.append(val)
        col_name = "_".join(parts) if parts else f"col_{col_idx}"
        combined_headers.append(col_name)
    
    # ‡∏ó‡∏≥‡πÉ‡∏´‡πâ column names unique
    seen = {}
    unique_headers = []
    for h in combined_headers:
        if h in seen:
            seen[h] += 1
            unique_headers.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 0
            unique_headers.append(h)
    
    df = pd.DataFrame(data[4:], columns=unique_headers)
    
    # ===== Clean Data =====
    df = df.dropna(axis=1, how='all')
    df = df.loc[:, ~df.columns.str.startswith('col_')]
    
    # ===== ‡πÉ‡∏´‡πâ GPT ‡∏ß‡∏¥‡πÄ‡∏Ñ‡∏£‡∏≤‡∏∞‡∏´‡πå‡πÇ‡∏Ñ‡∏£‡∏á‡∏™‡∏£‡πâ‡∏≤‡∏á‡∏Å‡πà‡∏≠‡∏ô =====
    print("ü§ñ GPT ‡∏Å‡∏≥‡∏•‡∏±‡∏á‡∏ß‡∏¥‡πÄ‡∏Ñ‡∏£‡∏≤‡∏∞‡∏´‡πå‡πÇ‡∏Ñ‡∏£‡∏á‡∏™‡∏£‡πâ‡∏≤‡∏á DataFrame...")
    structure = analyze_dataframe_structure(df)
    print(f"   üìå Month column: {structure.get('month_column')}")
    print(f"   üìå Type column: {structure.get('type_column')}")
    print(f"   üìå Type values: {structure.get('type_values')}")
    
    # ===== ‡πÉ‡∏´‡πâ GPT ‡∏™‡∏£‡πâ‡∏≤‡∏á column mapping =====
    print("\nü§ñ GPT ‡∏Å‡∏≥‡∏•‡∏±‡∏á‡πÅ‡∏õ‡∏•‡∏á‡∏ä‡∏∑‡πà‡∏≠ columns...")
    column_mapping = get_column_mapping_from_gpt(df.columns.tolist())
    print(f"   ‚úÖ ‡∏™‡∏£‡πâ‡∏≤‡∏á mapping ‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö {len(column_mapping)} columns")
    df = df.rename(columns=column_mapping)
    
    # Update structure ‡∏î‡πâ‡∏ß‡∏¢‡∏ä‡∏∑‡πà‡∏≠ column ‡πÉ‡∏´‡∏°‡πà
    if structure.get('month_column') and structure['month_column'] in column_mapping:
        structure['month_column'] = column_mapping[structure['month_column']]
    if structure.get('type_column') and structure['type_column'] in column_mapping:
        structure['type_column'] = column_mapping[structure['type_column']]
    
    # ===== ‡πÉ‡∏´‡πâ GPT ‡πÅ‡∏õ‡∏•‡∏á‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡πÉ‡∏ô‡πÅ‡∏ï‡πà‡∏•‡∏∞ cell =====
    thai_values = set()
    for col in df.columns:
        for val in df[col].dropna().unique():
            if isinstance(val, str) and any('\u0e00' <= c <= '\u0e7f' for c in val):
                thai_values.add(val)
    
    if thai_values:
        print(f"\nü§ñ GPT ‡∏Å‡∏≥‡∏•‡∏±‡∏á‡πÅ‡∏õ‡∏•‡∏á‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏• {len(thai_values)} ‡∏£‡∏≤‡∏¢‡∏Å‡∏≤‡∏£...")
        data_mapping = get_data_mapping_from_gpt(list(thai_values))
        print("   ‚úÖ ‡πÅ‡∏õ‡∏•‡∏á‡∏Ç‡πâ‡∏≠‡∏°‡∏π‡∏•‡πÄ‡∏£‡∏µ‡∏¢‡∏ö‡∏£‡πâ‡∏≠‡∏¢")
        
        for col in df.columns:
            df[col] = df[col].map(lambda x: data_mapping.get(x, x) if isinstance(x, str) else x)
        
        # Update type_values ‡∏î‡πâ‡∏ß‡∏¢‡∏Ñ‡πà‡∏≤‡πÉ‡∏´‡∏°‡πà
        if structure.get('type_values'):
            for key in structure['type_values']:
                old_val = structure['type_values'][key]
                if old_val and old_val in data_mapping:
                    structure['type_values'][key] = data_mapping[old_val]
    
    # ===== Filter rows ‡∏ï‡∏≤‡∏° structure ‡∏ó‡∏µ‡πà GPT ‡∏ß‡∏¥‡πÄ‡∏Ñ‡∏£‡∏≤‡∏∞‡∏´‡πå =====
    type_col = structure.get('type_column')
    type_vals = structure.get('type_values', {})
    
    if type_col and type_col in df.columns:
        valid_types = [v for v in type_vals.values() if v]
        if valid_types:
            print(f"\nüîç Filtering by {type_col} in {valid_types}")
            df = df[df[type_col].isin(valid_types)].copy()
    
    # Fill month ‡∏•‡∏á‡∏°‡∏≤
    month_col = structure.get('month_column')
    if month_col and month_col in df.columns:
        df[month_col] = df[month_col].ffill()
    
    df = df.reset_index(drop=True)
    
    # ===== ‡∏ï‡∏£‡∏ß‡∏à‡∏™‡∏≠‡∏ö‡∏ß‡πà‡∏≤‡∏¢‡∏±‡∏á‡∏°‡∏µ‡∏†‡∏≤‡∏©‡∏≤‡πÑ‡∏ó‡∏¢‡πÄ‡∏´‡∏•‡∏∑‡∏≠‡∏≠‡∏¢‡∏π‡πà‡πÑ‡∏´‡∏° =====
    remaining_thai = set()
    for col in df.columns:
        if any('\u0e00' <= c <= '\u0e7f' for c in col):
            remaining_thai.add(f"column: {col}")
        for val in df[col].dropna().unique():
            if isinstance(val, str) and any('\u0e00' <= c <= '\u0e7f' for c in val):
                remaining_thai.add(f"data: {val}")
    
    if remaining_thai:
        print(f"\n‚ö†Ô∏è ‡∏¢‡∏±‡∏á‡∏°‡∏µ‡∏†‡∏≤‡∏©‡∏≤‡πÑ‡∏ó‡∏¢‡πÄ‡∏´‡∏•‡∏∑‡∏≠: {remaining_thai}")
    
    return df


# ============ Main ============
if __name__ == "__main__":
    excel_path = "data/‡∏Ñ‡∏∏‡∏°‡∏é‡∏µ‡∏Å‡∏≤‡∏õ‡∏µ2567-‡∏£‡∏≤‡∏¢‡πÑ‡∏î‡πâ.xlsx"
    sheet_name = "‡∏™‡∏£‡∏∏‡∏õ‡∏£‡∏≤‡∏¢‡πÄ‡∏î‡∏∑‡∏≠‡∏ô 67"
    csv_path = "data/finance_2027_test.csv"
    
    print(f"üìä Converting: {excel_path}\n")
    
    df = convert_excel(excel_path, sheet_name)
    
    df.to_csv(csv_path, index=False, encoding='utf-8-sig')
    
    print(f"\n‚úÖ Saved to {csv_path}")
    print(f"   Rows: {len(df)}, Columns: {len(df.columns)}")
    
    print("\nüìã Columns:")
    for i, col in enumerate(df.columns, 1):
        print(f"  {i}. {col}")
    
    print("\nüìã Preview:")
    print(df.head())